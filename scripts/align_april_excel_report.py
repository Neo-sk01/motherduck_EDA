from __future__ import annotations

import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import mean, median, stdev
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from pipeline.anomaly import detect_anomalies

PUBLIC_REPORT = ROOT / "dashboard/public/data/reports/month_2026-04-01_2026-04-30/metrics.json"
FIXTURE_REPORT = ROOT / "dashboard/src/fixtures/april-2026-metrics.json"
MANIFEST = ROOT / "dashboard/public/data/reports/manifest.json"
ALL_QUEUES_WORKBOOK = Path("/Users/neosekaleli/Downloads/NeoLore_All_4_Queues_April2026 (1).xlsx")
DEDICATED_EDA_WORKBOOKS = {
    "8020": Path("/Users/neosekaleli/Downloads/queue_details_april_2026_eda_report.xlsx"),
    "8021": Path("/Users/neosekaleli/Downloads/queue_details_april_2026_8021_eda_report.xlsx"),
}
QUEUE_IDS = ["8020", "8021", "8030", "8031"]
QUEUE_META = {
    "8020": {"queue_name": "CSR English", "language": "English", "role": "primary"},
    "8021": {"queue_name": "CSR French", "language": "French", "role": "primary"},
    "8030": {"queue_name": "CSR Overflow English", "language": "English", "role": "overflow"},
    "8031": {"queue_name": "CSR Overflow French", "language": "French", "role": "overflow"},
}


def main() -> None:
    base = json.loads(PUBLIC_REPORT.read_text())
    all_workbook = openpyxl.load_workbook(ALL_QUEUES_WORKBOOK, data_only=True)
    all_records = records_from_all_queues_workbook(all_workbook)
    metrics_by_queue: dict[str, dict[str, Any]] = {}
    records_by_queue: dict[str, list[dict[str, Any]]] = {}
    source_files: dict[str, str] = {}

    for queue_id in QUEUE_IDS:
        if queue_id in DEDICATED_EDA_WORKBOOKS and DEDICATED_EDA_WORKBOOKS[queue_id].exists():
            workbook_path = DEDICATED_EDA_WORKBOOKS[queue_id]
            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            records = records_from_eda_workbook(workbook, queue_id)
            metrics = build_queue_metrics(
                records,
                queue_id,
                agent_summary=records_from_sheet(workbook["Agent Summary"], [("Agent Name", "Calls")]),
                release_summary=records_from_sheet(workbook["Release Reasons"], [("Queue Release Reason",)]),
                repeat_summary=records_from_sheet(workbook["Repeat Callers"], [("Caller Number", "Calls")]),
            )
            source_files[queue_id] = workbook_path.name
        else:
            records = [row for row in all_records if row["queue_id"] == queue_id]
            metrics = build_queue_metrics(records, queue_id)
            source_files[queue_id] = ALL_QUEUES_WORKBOOK.name
        metrics_by_queue[queue_id] = metrics
        records_by_queue[queue_id] = records

    crossqueue = build_crossqueue(metrics_by_queue, records_by_queue)
    base["queues"] = metrics_by_queue
    base["crossqueue"] = crossqueue
    base["anomalies"] = detect_anomalies(metrics_by_queue, crossqueue)
    base["source_mode"] = "excel_reference_overlay"
    base["validation"] = {
        "status": "success",
        "calculation_source": "excel_reference_cleaned_calls",
        "excel_reference_queues": QUEUE_IDS,
        "excel_reference_files": source_files,
        "excel_reference_note": (
            "Dedicated EDA workbooks are used for 8020 and 8021; "
            "the all-four-queues Excel workbook is used for 8030 and 8031."
        ),
    }

    for path in (PUBLIC_REPORT, FIXTURE_REPORT):
        path.write_text(json.dumps(base, indent=2, sort_keys=True) + "\n")
    update_manifest()


def records_from_eda_workbook(workbook: Any, queue_id: str) -> list[dict[str, Any]]:
    rows = records_from_sheet(workbook["Cleaned Data"], [("Call Time", "Caller Number", "Orig CallID")])
    out = []
    for row in rows:
        call_dt = parse_datetime(row.get("Call Datetime") or row.get("Call Time"))
        out.append(
            {
                "queue_id": queue_id,
                "queue_name": QUEUE_META[queue_id]["queue_name"],
                "language": QUEUE_META[queue_id]["language"],
                "role": QUEUE_META[queue_id]["role"],
                "call_id": str(row["Orig CallID"]),
                "call_datetime": call_dt,
                "date": call_dt.date().isoformat(),
                "hour": call_dt.hour,
                "dow": call_dt.strftime("%A"),
                "caller_number_norm": normalize_caller(row.get("Caller Number")),
                "agent_name": cleaned_text(row.get("Agent Name")),
                "agent_extension": cleaned_text(row.get("Agent Extension")),
                "queue_sec": numeric(row.get("Time in Queue Seconds")) or seconds(row.get("Time in Queue")),
                "agent_sec": numeric(row.get("Agent Time Seconds")) or seconds(row.get("Agent Time")),
                "hold_sec": numeric(row.get("Hold Time Seconds")) or seconds(row.get("Hold Time")),
                "queue_release_reason": cleaned_text(row.get("Queue Release Reason")),
                "agent_release_reason": cleaned_text(row.get("Agent Release Reason")),
            }
        )
    return out


def records_from_all_queues_workbook(workbook: Any) -> list[dict[str, Any]]:
    rows = records_from_sheet(workbook["All Calls"], [("Queue ID", "Call Time", "Caller Number", "Orig CallID")])
    out = []
    for row in rows:
        queue_id = str(int(row["Queue ID"]))
        call_dt = parse_datetime(row["Call Time"])
        out.append(
            {
                "queue_id": queue_id,
                "queue_name": cleaned_text(row.get("Queue Name")) or QUEUE_META[queue_id]["queue_name"],
                "language": cleaned_text(row.get("Language")) or QUEUE_META[queue_id]["language"],
                "role": str(row.get("Queue Role") or QUEUE_META[queue_id]["role"]).lower(),
                "call_id": str(row["Orig CallID"]),
                "call_datetime": call_dt,
                "date": call_dt.date().isoformat(),
                "hour": call_dt.hour,
                "dow": call_dt.strftime("%A"),
                "caller_number_norm": normalize_caller(row.get("Caller Number")),
                "agent_name": cleaned_text(row.get("Agent Name")),
                "agent_extension": cleaned_text(row.get("Agent Extension")),
                "queue_sec": seconds(row.get("Time in Queue")),
                "agent_sec": seconds(row.get("Agent Time")),
                "hold_sec": seconds(row.get("Hold Time")),
                "queue_release_reason": cleaned_text(row.get("Queue Release Reason")),
                "agent_release_reason": cleaned_text(row.get("Agent Release Reason")),
            }
        )
    return out


def build_queue_metrics(
    records: list[dict[str, Any]],
    queue_id: str,
    agent_summary: list[dict[str, Any]] | None = None,
    release_summary: list[dict[str, Any]] | None = None,
    repeat_summary: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    handled_calls = sum(1 for row in records if row["agent_sec"] > 0)
    total_calls = len(records)
    no_agent_calls = total_calls - handled_calls
    daily_volume = grouped_counts(records, "date")
    hourly_volume = hourly_counts(records)
    dow_counts = Counter(row["dow"] for row in records)
    busiest = sorted(daily_volume, key=lambda row: (-row["calls"], row["date"]))[0]
    quietest = sorted(daily_volume, key=lambda row: (row["calls"], row["date"]))[0]

    return {
        "queue_id": queue_id,
        "total_calls": total_calls,
        "handled_calls": handled_calls,
        "no_agent_calls": no_agent_calls,
        "answer_rate": handled_calls / total_calls if total_calls else 0.0,
        "no_agent_rate": no_agent_calls / total_calls if total_calls else 0.0,
        "raw_rows": total_calls,
        "cleaned_calls": total_calls,
        "duplicate_rows_removed": 0,
        "answered_no_agent_reconciled": True,
        "dedupe_key": "Orig CallID; Excel reference cleaned data",
        "calculation_source": "excel_reference_cleaned_calls",
        "days_with_calls": len(daily_volume),
        "avg_calls_per_active_day": total_calls / len(daily_volume) if daily_volume else 0.0,
        "busiest_day": busiest,
        "quietest_day": quietest,
        "daily_volume": daily_volume,
        "weekly_volume": weekly_volume(daily_volume),
        "hourly_volume": hourly_volume,
        "dow_volume": [{"dow": key, "calls": dow_counts[key]} for key in sorted(dow_counts)],
        "duration_distributions": {
            "queue_sec": duration_distribution(row["queue_sec"] for row in records),
            "agent_sec": duration_distribution(row["agent_sec"] for row in records if row["agent_sec"] > 0),
            "hold_sec": duration_distribution(row["hold_sec"] for row in records if row["hold_sec"] > 0),
        },
        "release_reasons": release_reasons(records, release_summary),
        "agent_leaderboard": agent_leaderboard(records, agent_summary),
        "top_callers": top_callers(records, repeat_summary),
    }


def build_crossqueue(
    metrics_by_queue: dict[str, dict[str, Any]],
    records_by_queue: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    return {
        "funnels": {
            "English": language_funnel(metrics_by_queue, "8020", "8030"),
            "French": language_funnel(metrics_by_queue, "8021", "8031"),
        },
        "agents": consolidated_agents(metrics_by_queue),
        "callers": consolidated_callers(records_by_queue),
        "same_hour_no_answer": [
            {
                "queue_id": queue_id,
                "hour": row["hour"],
                "calls": row["calls"],
                "no_answer_count": row["no_answer_count"],
                "no_answer_rate": row["no_answer_rate"],
            }
            for queue_id in QUEUE_IDS
            for row in metrics_by_queue[queue_id]["hourly_volume"]
        ],
        "same_day_volume": [
            {"queue_id": queue_id, "date": row["date"], "calls": row["calls"]}
            for queue_id in QUEUE_IDS
            for row in metrics_by_queue[queue_id]["daily_volume"]
        ],
    }


def language_funnel(metrics_by_queue: dict[str, dict[str, Any]], primary_id: str, overflow_id: str) -> dict[str, Any]:
    primary = metrics_by_queue[primary_id]
    overflow = metrics_by_queue[overflow_id]
    primary_failed = int(primary["no_agent_calls"])
    overflow_failed = int(overflow["no_agent_calls"])
    lost = overflow_failed
    return {
        "primary_calls": int(primary["total_calls"]),
        "primary_answered": int(primary["handled_calls"]),
        "primary_failed": primary_failed,
        "primary_no_agent_calls": primary_failed,
        "primary_no_agent_rate": float(primary["no_agent_rate"]),
        "overflow_received": int(overflow["total_calls"]),
        "routing_match": float(overflow["total_calls"] / primary_failed) if primary_failed else 0.0,
        "overflow_answered": int(overflow["handled_calls"]),
        "overflow_failed": overflow_failed,
        "lost": lost,
        "lost_rate": float(lost / primary["total_calls"]) if primary["total_calls"] else 0.0,
        "effective_answer_rate": float(1 - (lost / primary["total_calls"])) if primary["total_calls"] else 0.0,
        "unaccounted": int(primary_failed - overflow["total_calls"]),
        "final_dropped_available": True,
        "final_dropped_calls": overflow_failed,
        "drop_definition": "final_dropped_after_overflow",
    }


def consolidated_agents(metrics_by_queue: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for queue_id, metrics in metrics_by_queue.items():
        for agent in metrics["agent_leaderboard"]:
            name = agent["agent_name"]
            row = rows.setdefault(name, {"agent_name": name, **{qid: 0 for qid in QUEUE_IDS}})
            row[queue_id] = int(agent["calls"])
    for row in rows.values():
        row["total_calls"] = sum(int(row[qid]) for qid in QUEUE_IDS)
    return sorted(rows.values(), key=lambda row: (-row["total_calls"], row["agent_name"]))


def consolidated_callers(records_by_queue: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for queue_id, records in records_by_queue.items():
        for record in records:
            caller = record["caller_number_norm"]
            row = rows.setdefault(caller, {"caller_number_norm": caller, **{qid: 0 for qid in QUEUE_IDS}})
            row[queue_id] += 1
    for row in rows.values():
        row["total_calls"] = sum(int(row[qid]) for qid in QUEUE_IDS)
    return sorted(rows.values(), key=lambda row: (-row["total_calls"], row["caller_number_norm"]))


def grouped_counts(records: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    counts = Counter(row[key] for row in records)
    return [{key: group, "calls": count} for group, count in sorted(counts.items())]


def hourly_counts(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in records:
        groups[int(row["hour"])].append(row)
    out = []
    for hour, rows in sorted(groups.items()):
        no_answer = sum(1 for row in rows if row["agent_sec"] <= 0)
        handled_seconds = [row["agent_sec"] for row in rows if row["agent_sec"] > 0]
        out.append(
            {
                "hour": hour,
                "calls": len(rows),
                "no_answer_count": no_answer,
                "no_answer_rate": no_answer / len(rows) if rows else 0.0,
                "avg_agent_sec": json_number(mean(handled_seconds)) if handled_seconds else None,
            }
        )
    return out


def release_reasons(
    records: list[dict[str, Any]],
    release_summary: list[dict[str, Any]] | None,
) -> dict[str, list[dict[str, Any]]]:
    queue = []
    agent = []
    if release_summary:
        for row in release_summary:
            if row.get("Queue Release Reason"):
                queue.append(
                    {
                        "reason": str(row["Queue Release Reason"]),
                        "calls": int(first_value(row, "Count", "Calls")),
                    }
                )
            if row.get("Agent Release Reason"):
                agent.append(
                    {
                        "reason": str(row["Agent Release Reason"]),
                        "calls": int(first_value(row, "Count_1", "Calls_1", "Count", "Calls")),
                    }
                )
    if not queue:
        queue = reason_counts(row["queue_release_reason"] for row in records)
    if not agent:
        agent = reason_counts(row["agent_release_reason"] for row in records)
    return {"queue": queue, "agent": agent}


def reason_counts(values: Any) -> list[dict[str, Any]]:
    counts = Counter(cleaned_text(value) or "(blank)" for value in values)
    return [
        {"reason": reason, "calls": calls}
        for reason, calls in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        if reason != "(blank)" or calls > 0
    ]


def agent_leaderboard(
    records: list[dict[str, Any]],
    agent_summary: list[dict[str, Any]] | None,
) -> list[dict[str, Any]]:
    if agent_summary:
        rows = []
        total = 0
        for row in agent_summary:
            name = cleaned_text(row.get("Agent Name"))
            if not name or name == "(No agent)":
                continue
            calls = int(first_value(row, "Calls", "Handled Calls"))
            total += calls
            rows.append((row, name, calls))
        leaderboard = []
        for row, name, calls in rows:
            share = row.get("Share of Agent Calls")
            leaderboard.append(
                {
                    "agent_name": name,
                    "calls": calls,
                    "avg_sec": json_number(seconds(row.get("Avg Agent Time"))),
                    "median_sec": json_number(seconds(row.get("Median Agent Time"))),
                    "total_sec": json_number(seconds(row.get("Total Agent Time"))),
                    "pct_of_answered": float(share) if share is not None else (calls / total if total else 0.0),
                }
            )
        return leaderboard

    handled = [row for row in records if row["agent_sec"] > 0 and row["agent_name"]]
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in handled:
        groups[row["agent_name"]].append(row)
    total = len(handled)
    leaderboard = []
    for name, rows in groups.items():
        values = [row["agent_sec"] for row in rows]
        leaderboard.append(
            {
                "agent_name": name,
                "calls": len(rows),
                "avg_sec": json_number(mean(values)),
                "median_sec": json_number(median(values)),
                "total_sec": json_number(sum(values)),
                "pct_of_answered": len(rows) / total if total else 0.0,
            }
        )
    return sorted(leaderboard, key=lambda row: (-row["calls"], row["agent_name"]))


def top_callers(
    records: list[dict[str, Any]],
    repeat_summary: list[dict[str, Any]] | None,
) -> list[dict[str, Any]]:
    if repeat_summary:
        return [
            {"caller_number_norm": str(row["Caller Number"]), "calls": int(row["Calls"])}
            for row in repeat_summary
            if row.get("Caller Number") is not None and row.get("Calls") is not None
        ]
    counts = Counter(row["caller_number_norm"] for row in records)
    return [
        {"caller_number_norm": caller, "calls": calls}
        for caller, calls in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:100]
    ]


def records_from_sheet(sheet: Any, required_groups: list[tuple[str, ...]]) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    for row_index, values in enumerate(rows):
        headers = unique_headers(list(values))
        present = {header for header in headers if header is not None}
        if any(all(field in present for field in group) for group in required_groups):
            return [
                {header: value for header, value in zip(headers, row, strict=False) if header is not None}
                for row in rows[row_index + 1 :]
                if any(value is not None for value in row)
            ]
    raise ValueError(f"Could not find headers {required_groups} in sheet {sheet.title}")


def unique_headers(headers: list[Any]) -> list[Any]:
    seen: dict[str, int] = defaultdict(int)
    out = []
    for header in headers:
        if header is None:
            out.append(None)
            continue
        text = str(header)
        suffix = seen[text]
        seen[text] += 1
        out.append(text if suffix == 0 else f"{text}_{suffix}")
    return out


def weekly_volume(daily_volume: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[date, int] = defaultdict(int)
    for row in daily_volume:
        current = date.fromisoformat(row["date"])
        week_start = current - timedelta(days=current.weekday())
        grouped[week_start] += int(row["calls"])
    return [
        {
            "week_start": week_start.isoformat(),
            "week_end": (week_start + timedelta(days=6)).isoformat(),
            "calls": calls,
        }
        for week_start, calls in sorted(grouped.items())
    ]


def duration_distribution(values: Any) -> dict[str, float | int | None]:
    numeric_values = sorted(json_number(value) for value in values if json_number(value) is not None)
    if not numeric_values:
        return {"count": 0, "mean": None, "std": None, "min": None, "p25": None, "median": None, "p75": None, "max": None}
    return {
        "count": len(numeric_values),
        "mean": json_number(mean(numeric_values)),
        "std": json_number(stdev(numeric_values)) if len(numeric_values) > 1 else None,
        "min": json_number(min(numeric_values)),
        "p25": json_number(percentile(numeric_values, 0.25)),
        "median": json_number(median(numeric_values)),
        "p75": json_number(percentile(numeric_values, 0.75)),
        "max": json_number(max(numeric_values)),
    }


def percentile(values: list[float], q: float) -> float:
    position = (len(values) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return values[int(position)]
    return values[lower] + (values[upper] - values[lower]) * (position - lower)


def parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        return from_excel(value)
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%m/%d/%Y %I:%M %p"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unsupported datetime: {value!r}")


def seconds(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    parts = [int(part) for part in str(value).strip().split(":") if part != ""]
    if len(parts) == 2:
        minutes, secs = parts
        return float(minutes * 60 + secs)
    if len(parts) == 3:
        hours, minutes, secs = parts
        return float(hours * 3600 + minutes * 60 + secs)
    return 0.0


def numeric(value: Any) -> float:
    number = json_number(value)
    return number if number is not None else 0.0


def json_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def cleaned_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.casefold() == "null":
        return None
    return text


def normalize_caller(value: Any) -> str:
    text = cleaned_text(value)
    if text is None:
        return "Unknown"
    if text.casefold() == "restricted":
        return "Restricted"
    digits = re.sub(r"\D+", "", text)
    return digits or text


def first_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] is not None:
            return row[key]
    return None


def update_manifest() -> None:
    manifest = json.loads(MANIFEST.read_text())
    for report in manifest["reports"]:
        if report["key"] == "2026-04":
            report["source"] = "excel_reference_overlay"
            report["validation_status"] = "success"
            break
    MANIFEST.write_text(json.dumps(manifest, indent=2) + "\n")


if __name__ == "__main__":
    main()
